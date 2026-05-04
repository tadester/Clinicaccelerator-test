import argparse
import json
import re
import sys
from zipfile import BadZipFile
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter


HEADER_ROWS = {
    "section": 1,
    "label": 2,
    "focus": 3,
    "source": 4,
    "role": 5,
    "target_label": 6,
    "target": 7,
}
DATA_START_ROW = 8


class ConvertError(Exception):
    pass


def clean_text(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = re.sub(r"\s+", " ", value).strip()
        return value or None

    return value


def json_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time.min else value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return clean_text(value)


def make_id(text, col_letter):
    text = clean_text(text) or f"column {col_letter}"
    slug = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return f"{slug}_{col_letter.lower()}"


def merged_lookup(sheet):
    lookup = {}
    ranges = []

    for cell_range in sheet.merged_cells.ranges:
        anchor = sheet.cell(cell_range.min_row, cell_range.min_col).value
        anchor = clean_text(anchor)
        ranges.append(str(cell_range))

        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for col in range(cell_range.min_col, cell_range.max_col + 1):
                lookup[(row, col)] = anchor

    return lookup, ranges


def read_cell(sheet, values_sheet, row, col, merge_values):
    formula_cell = sheet.cell(row, col)
    value_cell = values_sheet.cell(row, col)

    value = None if isinstance(value_cell, MergedCell) else value_cell.value
    formula = None if isinstance(formula_cell, MergedCell) else formula_cell.value

    if value is None:
        value = merge_values.get((row, col))

    if isinstance(formula, str) and formula.startswith("="):
        return {
            "value": json_value(value),
            "formula": formula,
        }

    return json_value(value)


def column_has_data(sheet, col):
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        if clean_text(sheet.cell(row, col).value) is not None:
            return True
    return False


def build_columns(sheet, values_sheet, merge_values):
    columns = []

    for col in range(1, sheet.max_column + 1):
        letter = get_column_letter(col)
        label = read_cell(sheet, values_sheet, HEADER_ROWS["label"], col, merge_values)
        label_text = label.get("value") if isinstance(label, dict) else label
        has_data = column_has_data(values_sheet, col)

        if col == 1:
            kind = "date"
            metric_id = "week"
        elif label_text or has_data:
            kind = "metric"
            metric_id = make_id(label_text, letter)
        else:
            kind = "spacer"
            metric_id = f"spacer_{letter.lower()}"

        columns.append(
            {
                "id": metric_id,
                "kind": kind,
                "column": letter,
                "index": col,
                "section": read_cell(sheet, values_sheet, HEADER_ROWS["section"], col, merge_values),
                "label": label,
                "focus": read_cell(sheet, values_sheet, HEADER_ROWS["focus"], col, merge_values),
                "source": read_cell(sheet, values_sheet, HEADER_ROWS["source"], col, merge_values),
                "role": read_cell(sheet, values_sheet, HEADER_ROWS["role"], col, merge_values),
                "target_label": read_cell(sheet, values_sheet, HEADER_ROWS["target_label"], col, merge_values),
                "target": read_cell(sheet, values_sheet, HEADER_ROWS["target"], col, merge_values),
            }
        )

    return columns


def build_records(sheet, values_sheet, columns, merge_values):
    records = []
    metric_columns = [col for col in columns if col["kind"] == "metric"]

    for row in range(DATA_START_ROW, sheet.max_row + 1):
        week = read_cell(sheet, values_sheet, row, 1, merge_values)
        if week is None:
            continue

        values = {}
        for col_info in metric_columns:
            col = col_info["index"]
            value = read_cell(sheet, values_sheet, row, col, merge_values)
            if value is not None:
                values[col_info["id"]] = {
                    "column": col_info["column"],
                    "label": col_info["label"],
                    "value": value,
                }

        records.append(
            {
                "row": row,
                "week": week,
                "values": values,
            }
        )

    return records


def convert(input_path):
    input_path = Path(input_path)

    if not input_path.exists():
        raise ConvertError(f"Input file was not found: {input_path}")

    if not input_path.is_file():
        raise ConvertError(f"Input path is not a file: {input_path}")

    try:
        workbook = load_workbook(input_path, data_only=False)
        values_workbook = load_workbook(input_path, data_only=True)
    except (InvalidFileException, BadZipFile, OSError) as error:
        raise ConvertError(f"Could not read Excel workbook: {error}") from error

    if not workbook.worksheets:
        raise ConvertError("Workbook does not contain any sheets.")

    output = {
        "source_file": input_path.name,
        "sheets": [],
    }

    for sheet in workbook.worksheets:
        values_sheet = values_workbook[sheet.title]
        merge_values, ranges = merged_lookup(sheet)
        columns = build_columns(sheet, values_sheet, merge_values)

        output["sheets"].append(
            {
                "name": sheet.title,
                "size": {
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                },
                "merged_ranges": ranges,
                "header_rows": HEADER_ROWS,
                "data_start_row": DATA_START_ROW,
                "columns": columns,
                "records": build_records(sheet, values_sheet, columns, merge_values),
            }
        )

    return output


def main():
    parser = argparse.ArgumentParser(description="Convert the clinic scoreboard spreadsheet to JSON.")
    parser.add_argument("input", nargs="?", default="Scoreboard Test.xlsx")
    parser.add_argument("output", nargs="?", default="output.json")
    args = parser.parse_args()

    try:
        data = convert(args.input)
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    except ConvertError as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1
    except OSError as error:
        print(f"Error: could not write output file: {error}", file=sys.stderr)
        return 1

    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
